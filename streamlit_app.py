import streamlit as st
import pandas as pd
from shapely.geometry import Point, Polygon
import utm
import io
from datetime import datetime
import json
from shapely.geometry import Polygon
from collections import Counter
from openpyxl import load_workbook

def load_boundary_polygon_from_geojson():
    geojson_str = """
    {
      "type": "FeatureCollection",
      "features": [
        {
          "type": "Feature",
          "properties": {},
          "geometry": {
            "type": "Polygon",
            "coordinates": [[
              [34.244109701965954, 31.34515124626691],
              [34.271801013150906, 31.323114353523096],
              [34.27221922502832, 31.323347944447406],
              [34.27429699743249, 31.325468796399363],
              [34.2761950349535, 31.324465748123643],
              [34.28010470302689, 31.320070567019116],
              [34.28209500652065, 31.318297579264524],
              [34.28172504986043, 31.317912820325915],
              [34.28458819271049, 31.31659363488754],
              [34.2848133837218, 31.317280712942775],
              [34.2894881813163, 31.316717335594504],
              [34.29017983942086, 31.316387536780937],
              [34.29199623276878, 31.317581497651048],
              [34.29582053057089, 31.319316986520263],
              [34.29811894182947, 31.32191102969746],
              [34.303491313117746, 31.327022500349045],
              [34.303410887756655, 31.325882066033415],
              [34.30450467266664, 31.325524818706853],
              [34.30960225029642, 31.32993107828581],
              [34.31061945042602, 31.33099090481408],
              [34.310639282219825, 31.330987357960282],
              [34.31108205888856, 31.330618072535586],
              [34.311693291632025, 31.330920342467223],
              [34.31193456771524, 31.3292028867629],
              [34.31279621564519, 31.327167215969283],
              [34.31649822922131, 31.329309867908165],
              [34.31660640922044, 31.329199690752304],
              [34.317800548964385, 31.33030857005474],
              [34.31940122756521, 31.33131229011842],
              [34.3216939268211, 31.33399803218731],
              [34.32248797348197, 31.33348815180163],
              [34.322501848306985, 31.33420356326917],
              [34.322799550792524, 31.337306508650755],
              [34.32260060091363, 31.339099961837604],
              [34.32229933748644, 31.339701225887595],
              [34.32180038752463, 31.340403206330137],
              [34.32169894167748, 31.341004462051757],
              [34.32209857683557, 31.341206630143148],
              [34.322200720823076, 31.341498760325806],
              [34.32289925459685, 31.34230510609541],
              [34.32259935811584, 31.34300520333032],
              [34.323095649541216, 31.345974652641843],
              [34.32458958809403, 31.349408716239296],
              [34.324902624386766, 31.35081033647745],
              [34.322806301772545, 31.352602759884093],
              [34.33240644928105, 31.361557028466734],
              [34.333803339883644, 31.363515855319633],
              [34.33650746777644, 31.36874984661597],
              [34.342823707403056, 31.37424255802813],
              [34.36263048463758, 31.395226046740675],
              [34.36804897270062, 31.4001357354432],
              [34.37029972927985, 31.398890548844193],
              [34.37109829806835, 31.40171005067063],
              [34.37389944051748, 31.408147869404246],
              [34.38129121604263, 31.418523223409892],
              [34.38200386917157, 31.420228535702904],
              [34.37656367378668, 31.423630385915686],
              [34.378796372367674, 31.42544844210188],
              [34.37750876733048, 31.426309019811],
              [34.37849692154387, 31.427493467748207],
              [34.37309328001248, 31.431006083104236],
              [34.375102202210286, 31.432102628835395],
              [34.37149795944467, 31.435997153103045],
              [34.36829254682138, 31.43408142028568],
              [34.36390936867167, 31.438394440400813],
              [34.366597779259365, 31.442124870210293],
              [34.36389459718518, 31.443511140740526],
              [34.36068918456181, 31.44392701790055],
              [34.3564202248929, 31.443385116994108],
              [34.355785050962595, 31.44392701790055],
              [34.310564322422294, 31.401768178909634],
              [34.29773075205722, 31.392930696134982],
              [34.28819110733389, 31.384407129065693],
              [34.26731091610063, 31.365103525604297],
              [34.244109701965954, 31.34515124626691]
            ]]
          }
        }
      ]
    }
    """

    geojson_data = json.loads(geojson_str)
    coordinates = geojson_data['features'][0]['geometry']['coordinates'][0]

    # Coordinates are in [lon, lat], but shapely expects (lon, lat) for Point and Polygon
    polygon = Polygon([(lon, lat) for lon, lat in coordinates])
    return polygon

boundary_polygon = load_boundary_polygon_from_geojson()

def convert_to_utm(lat, lon):
    easting, northing, zone, _ = utm.from_latlon(lat, lon)
    return round(easting), round(northing)

def check_within_boundary(lat, lon):
    return boundary_polygon.contains(Point(lon, lat))

def deduplicate_headers(columns):
    counts = Counter()
    new_cols = []
    for col in columns:
        counts[col] += 1
        new_cols.append(f"{col} {counts[col]}" if counts[col] > 1 else col)
    return new_cols

def process_excel(uploaded_file):
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook

    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file)
    sheet = workbook.active

    # Detect and delete hidden columns
    hidden_cols = []
    for i in range(1, 4):
        col_letter = get_column_letter(i)
        if sheet.column_dimensions[col_letter].hidden:
            hidden_cols.append(i)

    keep_format = len(hidden_cols) == 0

    for i in reversed(hidden_cols):
        sheet.delete_cols(i)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    df = pd.read_excel(buffer)
    df.columns = [str(col).strip().replace('\xa0', ' ').lower() for col in df.columns]
    df.columns = deduplicate_headers(df.columns)

    if 'organization name' not in df.columns:
        raise ValueError("Missing 'organization name' column")

    lat_cols = [col for col in df.columns if col.startswith('lat')]
    lon_cols = [col for col in df.columns if col.startswith('log')]
    point_cols = [col for col in df.columns if col.startswith('point')]

    df.insert(0, 'CLA SN', '')
    df.insert(1, 'CLA Status', '')
    df.insert(2, 'Not recommended', '')

    for idx, row in df.iterrows():
        if pd.isna(row.get("organization name")) or str(row["organization name"]).strip() == "":
            continue

        total = 0
        outside = []

        for i in range(len(lat_cols)):
            lat = row.get(lat_cols[i]) if i < len(lat_cols) else None
            lon = row.get(lon_cols[i]) if i < len(lon_cols) else None
            point_name = row.get(point_cols[i]) if i < len(point_cols) else None

            if pd.isna(lat) and pd.isna(lon) and (pd.isna(point_name) or str(point_name).strip() == ''):
                continue

            if pd.notna(lat) and pd.notna(lon):
                try:
                    lat = float(lat)
                    lon = float(lon)
                    total += 1
                    if not check_within_boundary(lat, lon):
                        outside.append(f"Point {i + 1}")
                except:
                    outside.append(f"Point {i + 1}")
            else:
                outside.append(f"Point {i + 1}")

        if total == 0:
            status = ""
        elif len(outside) == total:
            status = "Not recommended"
        elif len(outside) == 0:
            status = "Acknowledged"
        else:
            status = "Partially acknowledged"

        df.at[idx, 'CLA Status'] = status
        df.at[idx, 'Not recommended'] = "All points" if len(outside) == total else ", ".join(outside)

    return df, keep_format

def generate_ab_text(df):
    lat_cols = [col for col in df.columns if col.startswith('lat')]
    lon_cols = [col for col in df.columns if col.startswith('log')]
    point_cols = [col for col in df.columns if col.startswith('point')]

    text = ""
    char_count = 0
    cap_marker = "\n----------------------------5000 LETTERS CAP------------------------------\n\n"

    for _, row in df.iterrows():
        section = ""
        org = str(row.get('organization name', '')).split(' - ')[0]

        # Format date and time fields cleanly
        def format_dt(value, fmt):
            if pd.isna(value) or value == '':
                return ''
            if isinstance(value, datetime):
                return value.strftime(fmt)
            try:
                return pd.to_datetime(value).strftime(fmt)
            except:
                return str(value)

        date_str = format_dt(row.get('date', ''), '%Y-%m-%d')
        start_str = format_dt(row.get('start time', ''), '%H:%M')
        end_str = format_dt(row.get('end time', ''), '%H:%M')
        purpose = str(row.get('purpose', '')).strip()

        section += f"{org}\n"
        section += f"{date_str}  {start_str}-{end_str}\n"
        section += f"{purpose}\n"

        for i, (lat_col, lon_col) in enumerate(zip(lat_cols, lon_cols)):
            lat = row.get(lat_col)
            lon = row.get(lon_col)
            point_name = row.get(point_cols[i]) if i < len(point_cols) else f"Point {i+1}"

            if pd.isna(lat) or pd.isna(lon):
                continue

            try:
                lat = float(lat)
                lon = float(lon)
                easting, northing = convert_to_utm(lat, lon)

                # Custom formatting: remove leading '3' from northing
                northing_str = str(northing)
                if northing_str.startswith("3"):
                    northing_str = northing_str[1:]

                section += f"{point_name}\n{easting}, {northing_str}\n"
            except:
                continue

        section += "\n"
        if char_count + len(section) > 4500:
            text += cap_marker
            char_count = 0

        text += section
        char_count += len(section)

    return text

# Streamlit app
st.title("Notifications App")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx", "xls"])

if uploaded_file:
    try:
        df_processed, keep_format = process_excel(uploaded_file)
        st.success("Excel file processed successfully!")

        base_name = uploaded_file.name.rsplit('.', 1)[0]
        excel_buffer = io.BytesIO()

        if keep_format:
            # Preserve original formatting and insert data using openpyxl
            uploaded_file.seek(0)
            workbook = load_workbook(uploaded_file)
            sheet = workbook.active

            sheet.insert_cols(1, amount=3)
            sheet.cell(row=1, column=1).value = "CLA SN"
            sheet.cell(row=1, column=2).value = "CLA Status"
            sheet.cell(row=1, column=3).value = "Not recommended"

            for i, row in df_processed.iterrows():
                sheet.cell(row=i + 2, column=2).value = row['CLA Status']
                sheet.cell(row=i + 2, column=3).value = row['Not recommended']

            workbook.save(excel_buffer)
            excel_buffer.seek(0)

        else:
            # Use Pandas to write clean file (e.g., when hidden columns were deleted)
            with pd.ExcelWriter(excel_buffer, engine='openpyxl', datetime_format='DD/MM/YYYY', date_format='DD/MM/YYYY') as writer:
                df_processed.to_excel(writer, index=False)
            excel_buffer.seek(0)

        processed_excel_filename = f"{base_name} - Processed Coordinates.xlsx"
        st.download_button("Download Processed Excel", excel_buffer.getvalue(), file_name=processed_excel_filename)

        ab_text = generate_ab_text(df_processed)
        ab_text_filename = f"{base_name} - AB Output.txt"
        st.download_button("Download AB Text File", ab_text, file_name=ab_text_filename)

        st.dataframe(df_processed.head())

    except Exception as e:
        st.error(f"Error processing file: {e}")
